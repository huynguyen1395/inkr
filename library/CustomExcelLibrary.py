import openpyxl
from pathlib import Path
import sys


class CustomExcelLibrary:
    @staticmethod
    def write_data_to_excel(excel_path, sheet_name, data_lists):
        """
        Writes data from multiple lists into an Excel file.
        
        :param sheet_name:
        :param excel_path: Path to the Excel file.
        :param data_lists: List of lists containing data to write.
        """
        book = openpyxl.load_workbook(excel_path)
        # sheet = book.active
        sheet = book.create_sheet(title=sheet_name)
        # sheet = book[sheet_name]
        for row_idx, data_list in enumerate(data_lists, start=1):
            for col_idx, value in enumerate(data_list, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)
        book.save(excel_path)

    @staticmethod
    def count_keyword(excel_path, keyword):
        # Load the workbook
        workbook = openpyxl.load_workbook(excel_path)

        # Initialize counter
        keyword_count = 0

        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Iterate through all cells in the sheet
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and keyword in str(cell.value):
                        keyword_count += 1
        return keyword_count

    @staticmethod
    def find_most_common_keywords(excel_path, num_keywords=100, exclude_words=None):
        """
        Finds the most frequently appearing keywords in an Excel file.

        :param excel_path: Path to the Excel file
        :param num_keywords: Number of top keywords to return (default: 20)
        :param exclude_words: Optional list of common words to exclude (e.g., 'the', 'and', 'a')
        :return: List of tuples containing (keyword, count) sorted by count in descending order
        """
        # Default exclude list if none provided
        if exclude_words is None:
            exclude_words = [
                "the", "and", "a", "to", "of", "in", "is", "that", "it", "for",
                "on", "with", "as", "at", "this", "by", "be", "or", "an", "are",
                "was", "were", "from", "have", "has", "had", "not", "but", "what",
                "all", "when", "who", "which", "will", "there", "can", "more", "no",
                "if", "so", "their", "would", "about", "up", "out", "them", "then",
                "some", "these", "his", "her", "they", "could", "into", "only", "one",
                "been", "other", "do", "you", "your", "my", "we", "us", "our", "me"
            ]

        # Load the workbook
        workbook = openpyxl.load_workbook(excel_path)

        # Initialize word counter dictionary
        word_counts = {}

        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Iterate through all cells in the sheet
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Split cell text into words
                        words = str(cell.value).lower().split()

                        # Count each word
                        for word in words:
                            # Remove any punctuation
                            word = word.strip('.,!?:;()"\'')

                            # Skip empty strings or excluded words
                            if word and word not in exclude_words:
                                if word in word_counts:
                                    word_counts[word] += 1
                                else:
                                    word_counts[word] = 1

        # Sort words by count in descending order and get the top N
        if not word_counts:
            return []

        # Sort by count (descending) and get top num_keywords
        top_keywords = sorted(word_counts.items(), key=lambda x: x[1], reverse=True)[:num_keywords]
        return top_keywords

    @staticmethod
    def generate_keyword_chart(top_keywords, chart_title="Top Keywords", output_path=None):
        """
        Generates a bar chart of the top keywords using matplotlib.

        :param top_keywords: List of tuples (keyword, count) from find_most_common_keywords
        :param chart_title: Title for the chart
        :param output_path: Path to save the chart image (optional)
        """
        try:
            import matplotlib.pyplot as plt

            # Unpack keywords and counts
            keywords = [item[0] for item in top_keywords]
            counts = [item[1] for item in top_keywords]

            # Create figure with appropriate size based on number of keywords
            plt.figure(figsize=(10, max(6, len(keywords) * 0.3)))

            # Create horizontal bar chart (easier to read for text labels)
            bars = plt.barh(keywords, counts, color='skyblue')

            # Add count labels on the bars
            for bar in bars:
                width = bar.get_width()
                plt.text(width + 0.3, bar.get_y() + bar.get_height() / 2,
                         f'{width:.0f}', ha='left', va='center')

            # Add labels and title
            plt.xlabel('Occurrences')
            plt.ylabel('Keywords')
            plt.title(chart_title)
            plt.tight_layout()

            # Save the chart if output path is provided
            if output_path:
                plt.savefig(output_path)
                print(f"Chart saved to {output_path}")

            # Display the chart
            plt.show()
        except ImportError:
            print("Matplotlib is not installed. Please install it to generate charts.")

    @staticmethod
    def export_keywords_to_excel_with_chart(top_keywords, output_excel_path, sheet_name="cnn"):
        """
        Exports the top keywords to a new Excel file with a chart.

        :param top_keywords: List of tuples (keyword, count) from find_most_common_keywords
        :param output_excel_path: Path to save the new Excel file
        :param sheet_name: Name of the sheet for the data and chart
        """
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference

        # Create a new workbook and get active sheet
        # wb = Workbook()
        # ws = wb.active
        wb = openpyxl.load_workbook(output_excel_path)
        ws = wb.active
        ws.title = sheet_name

        # Add headers
        ws['A1'] = "Keyword"
        ws['B1'] = "Occurrences"

        # Add data
        for i, (keyword, count) in enumerate(top_keywords, start=2):
            ws.cell(row=i, column=1, value=keyword)
            ws.cell(row=i, column=2, value=count)

        # Create chart
        chart = BarChart()
        chart.type = "bar"  # Horizontal bar chart
        chart.title = "Top Keywords by Frequency"
        chart.y_axis.title = "Keywords"
        chart.x_axis.title = "Occurrences"

        # Add data to chart
        data = Reference(ws, min_col=2, min_row=1, max_row=len(top_keywords) + 1, max_col=2)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(top_keywords) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Add chart to worksheet (position it below the data)
        ws.add_chart(chart, f"D2")

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook
        wb.save(output_excel_path)
        return f"Excel file with chart created at: {output_excel_path}"

    @staticmethod
    def save_keywords_to_existing_excel(excel_path, top_keywords, sheet_name, add_chart=True):
        """
        Saves the top keywords to a new worksheet in an existing Excel file with formatting.

        :param excel_path: Path to the existing Excel file
        :param top_keywords: List of tuples (keyword, count) from find_most_common_keywords
        :param sheet_name: Name of the new worksheet to create (or use if exists)
        :param add_chart: Whether to add a bar chart to visualize the data
        :return: Status message
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, Reference
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            print("Debug message", sheet_name, excel_path)
            # Load the existing workbook
            try:
                workbook = openpyxl.load_workbook(excel_path)
            except FileNotFoundError:
                return f"Error: Excel file not found at {excel_path}"

            # Check if the sheet already exists
            if sheet_name in workbook.sheetnames:
                # Use the existing sheet
                worksheet = workbook[sheet_name]
                # Clear existing content
                for row in worksheet.rows:
                    for cell in row:
                        cell.value = None
            else:
                # Create a new worksheet
                worksheet = workbook.create_sheet(title=sheet_name)
            print("Debug message", sheet_name, excel_path)

            # Define styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Add title
            title_cell = worksheet['A1']
            title_cell.value = "Keyword Analysis Results"
            title_cell.font = Font(bold=True, size=14)
            worksheet.merge_cells('A1:B1')
            title_cell.alignment = Alignment(horizontal='center')

            # Add headers
            worksheet['A2'] = "Keyword"
            worksheet['B2'] = "Occurrences"

            # Apply header styling
            for col in range(1, 3):
                cell = worksheet.cell(row=2, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Add data
            for i, (keyword, count) in enumerate(top_keywords, start=3):
                # Add keyword
                keyword_cell = worksheet.cell(row=i, column=1, value=keyword)
                keyword_cell.border = border

                # Add count
                count_cell = worksheet.cell(row=i, column=2, value=count)
                count_cell.border = border
                count_cell.alignment = Alignment(horizontal='center')

                # Add alternating row colors
                if i % 2 == 0:
                    for col in range(1, 3):
                        worksheet.cell(row=i, column=col).fill = PatternFill(
                            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                        )

            # Add a chart if requested
            if add_chart:
                # Create chart
                chart = BarChart()
                chart.type = "bar"  # Horizontal bar chart
                chart.title = "Top Keywords by Frequency"
                chart.y_axis.title = "Keywords"
                chart.x_axis.title = "Occurrences"
                chart.style = 10  # A clean chart style

                # Define the data range for the chart
                data = Reference(worksheet, min_col=2, min_row=2, max_row=len(top_keywords) + 2, max_col=2)
                categories = Reference(worksheet, min_col=1, min_row=3, max_row=len(top_keywords) + 2)

                # Add data to chart
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)

                # Add chart to worksheet - position it to the right of the data
                chart_cell = f"D3"
                worksheet.add_chart(chart, chart_cell)

            # Auto-adjust column widths
            for col in range(1, 3):
                max_length = 0
                column_letter = get_column_letter(col)
                for row in range(1, len(top_keywords) + 3):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 4
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            workbook.save(excel_path)

            return f"Keywords saved to worksheet '{sheet_name}' in {excel_path} with formatting"

        except Exception as e:
            return f"Error saving keywords to Excel: {str(e)}"

    def get_account(self, excel_path, role):
        book = openpyxl.load_workbook(excel_path)
        sheet = book.active
        role = self.check_role(role)
        row = role + 1
        row_data = []
        for cell in sheet[row]:
            row_data.append(cell.value)
        return row_data

    @staticmethod
    def get_all_rows_except_first(excel_path, user):
        book = openpyxl.load_workbook(excel_path)
        sheet = book.active
        all_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if user in str(row):
                all_rows.append(list(row))
        return all_rows[0]

    @staticmethod
    def check_role(value):
        switch_dict = {
            "admin": 1,
            "hr": 2,
            "employee": 3
        }
        return switch_dict.get(value, 1)

    @staticmethod
    def get_account_by_uid(excel_path, uid):
        book = openpyxl.load_workbook(excel_path)
        sheet = book.active
        row = str(uid).split("_", 2)
        row = int(row[1])
        row_data = []
        for cell in sheet[row]:
            row_data.append(cell.value)
        return row_data

    @staticmethod
    def split_list(input_list, size):
        """
        Splits a single list into multiple smaller lists of a given size.
    
        :param input_list: List to be split
        :param size: Size of each smaller list
        :return: List of smaller lists
        """
        return [input_list[i:i + int(size)] for i in range(0, len(input_list), int(size))]



    @staticmethod
    def write_api_response_to_excel(excel_path, sheet_name, response_data):
        """
        Writes API response data to an Excel file with all keys as headers and values as rows.
        Specifically handles 'quote' dictionary by separating it into multiple columns.

        :param excel_path: Path to the Excel file
        :param sheet_name: Name of the sheet to write data to
        :param response_data: API response data (list of dictionaries)
        :return: Status message
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            # Load or create the workbook
            try:
                workbook = openpyxl.load_workbook(excel_path)
            except FileNotFoundError:
                workbook = openpyxl.Workbook()

            # Check if the sheet already exists
            # if sheet_name in workbook.sheetnames:
            #     # Use the existing sheet
            #     worksheet = workbook[sheet_name]
            #     # Clear existing content
            #     for row in worksheet.rows:
            #         for cell in row:
            #             cell.value = None
            # else:
                # Create a new worksheet
            worksheet = workbook.create_sheet(title=sheet_name)

            # Define styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Verify response_data is a list of dictionaries
            if not isinstance(response_data, list):
                worksheet.cell(row=1, column=1, value="Error: Expected a list of dictionaries")
                workbook.save(excel_path)
                return "Error: Expected a list of dictionaries for response_data"

            if not response_data:
                worksheet.cell(row=1, column=1, value="No data in response")
                workbook.save(excel_path)
                return "No data in response"

            # Process data to collect all column headers, expanding nested dictionaries
            all_headers = []
            flattened_data = []

            for item in response_data:
                if not isinstance(item, dict):
                    continue

                flattened_item = {}

                for key, value in item.items():
                    # Special handling for 'quote' dictionary
                    if key == 'quote' and isinstance(value, dict):
                        for currency, quote_data in value.items():
                            if isinstance(quote_data, dict):
                                for quote_key, quote_value in quote_data.items():
                                    # Create a combined header for quote fields with format: quote_USD_price
                                    flat_key = f"quote_{currency}_{quote_key}"
                                    flattened_item[flat_key] = quote_value
                                    if flat_key not in all_headers:
                                        all_headers.append(flat_key)
                    elif isinstance(value, dict):
                        # Handle other nested dictionaries by flattening them
                        for nested_key, nested_value in value.items():
                            flat_key = f"{key}_{nested_key}"
                            flattened_item[flat_key] = nested_value
                            if flat_key not in all_headers:
                                all_headers.append(flat_key)
                    else:
                        # Regular key-value pair
                        flattened_item[key] = value
                        if key not in all_headers:
                            all_headers.append(key)

                flattened_data.append(flattened_item)

            # Sort headers for consistent column order, but prioritize common important fields
            priority_fields = ['id', 'name', 'symbol', 'slug', 'rank', 'is_active', 'last_updated']

            # Custom sort function to prioritize certain fields
            def custom_sort(header):
                if header in priority_fields:
                    return (0, priority_fields.index(header))
                elif header.startswith('quote_USD_'):
                    return (1, header)
                else:
                    return (2, header)

            all_headers.sort(key=custom_sort)

            # Write headers
            for col_idx, header in enumerate(all_headers, start=1):
                cell = worksheet.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Write data rows
            for row_idx, item in enumerate(flattened_data, start=2):
                for col_idx, header in enumerate(all_headers, start=1):
                    value = item.get(header, "")

                    # Handle any remaining nested structures
                    if isinstance(value, (dict, list)):
                        value = str(value)

                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border

                    # Format numeric values
                    if isinstance(value, (int, float)):
                        if header.endswith('_price') or header.endswith('_market_cap') or header.endswith('_volume_24h'):
                            cell.number_format = '$#,##0.00'
                        elif header.endswith('_percent_change_1h') or header.endswith('_percent_change_24h') or header.endswith('_percent_change_7d'):
                            cell.number_format = '0.00%'
                            # Apply color formatting for percent changes
                            if value > 0:
                                cell.font = Font(color="00AA00")  # Green for positive
                            elif value < 0:
                                cell.font = Font(color="DD0000")  # Red for negative

                    # Add alternating row colors
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Auto-adjust column widths for better readability
            for col_idx, header in enumerate(all_headers, start=1):
                max_length = len(str(header))
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter

                for row_idx in range(2, min(len(flattened_data) + 2, 30)):  # Check only the first 30 rows for width
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, min(len(str(cell_value)), 50))  # Limit to 50 chars

                adjusted_width = max_length + 3
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            workbook.save(excel_path)

            return f"API response data written to worksheet '{sheet_name}' with expanded quote fields in {excel_path}"

        except Exception as e:
            return f"Error writing API response to Excel: {str(e)}"

    @staticmethod
    def clean_coin_market_data(excel_path, sheet_name=None, decimal_places=2):
        """
        Cleans the coin market data Excel file by:
        1. Removing 'quote_USD_' prefix from column headers
        2. Converting formatted values (with $, %) to plain float numbers
        3. Formatting all numeric values to have a consistent number of decimal places

        :param excel_path: Path to the Excel file (coin_market_data.xlsx)
        :param sheet_name: Name of the sheet to clean (None for all sheets)
        :param decimal_places: Number of decimal places for numeric values (default: 2)
        :return: Status message
        """
        try:
            import openpyxl
            import re

            # Load the workbook
            workbook = openpyxl.load_workbook(excel_path)

            # Determine which sheets to process
            sheets_to_process = []
            if sheet_name is None:
                # Process all sheets
                sheets_to_process = workbook.sheetnames
            else:
                # Process only the specified sheet
                if sheet_name in workbook.sheetnames:
                    sheets_to_process = [sheet_name]
                else:
                    return f"Error: Sheet '{sheet_name}' not found in {excel_path}"

            # Track changes for the status message
            cleaned_sheets = 0
            cleaned_headers = 0
            cleaned_values = 0
            formatted_numbers = 0

            # Process each sheet
            for current_sheet_name in sheets_to_process:
                worksheet = workbook[current_sheet_name]

                # Get all headers from the first row
                headers = []
                for cell in worksheet[1]:
                    headers.append(cell.value)

                # Track if any changes were made to this sheet
                sheet_changed = False

                # Clean headers - remove 'quote_USD_' prefix
                for col_idx, header in enumerate(headers, start=1):
                    if header and isinstance(header, str) and 'quote_USD_' in header:
                        # Remove the 'quote_USD_' prefix
                        new_header = header.replace('quote_USD_', '')
                        # Update the cell
                        worksheet.cell(row=1, column=col_idx, value=new_header)
                        cleaned_headers += 1
                        sheet_changed = True

                # Clean values - convert formatted numbers to plain float values
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    for col_idx, cell in enumerate(row, start=1):
                        if cell.value is not None:
                            value_changed = False

                            # Convert string values with $ or % to float
                            if isinstance(cell.value, str):
                                # Check for $ format
                                if '$' in cell.value:
                                    try:
                                        # Remove $ and commas, then convert to float
                                        clean_value = cell.value.replace('$', '').replace(',', '')
                                        cell.value = float(clean_value)
                                        value_changed = True
                                    except ValueError:
                                        pass

                                # Check for % format
                                elif '%' in cell.value:
                                    try:
                                        # Remove % and convert to decimal float
                                        clean_value = cell.value.replace('%', '')
                                        cell.value = float(clean_value) / 100
                                        value_changed = True
                                    except ValueError:
                                        pass

                            # Format numeric values to specified decimal places
                            if isinstance(cell.value, (int, float)):
                                # Round to specified decimal places
                                original_value = cell.value
                                cell.value = round(cell.value, decimal_places)

                                # Ensure values are exactly as specified (e.g., 1.00 instead of 1)
                                # by setting the number format
                                if isinstance(cell.value, float):
                                    # Set format to display exactly the specified number of decimal places
                                    cell.number_format = f'0.{"0" * decimal_places}'
                                    formatted_numbers += 1
                                    sheet_changed = True

                            # Check if the cell has a number format that we should clear
                            if cell.number_format and ('$' in cell.number_format or '%' in cell.number_format):
                                # For numeric values, update the formatting
                                if isinstance(cell.value, (int, float)):
                                    # Set to general number format with specified decimal places
                                    cell.number_format = f'0.{"0" * decimal_places}'
                                    value_changed = True

                            if value_changed:
                                cleaned_values += 1
                                sheet_changed = True

                if sheet_changed:
                    cleaned_sheets += 1

            # Save the workbook with the changes
            workbook.save(excel_path)

            return (f"Cleaned {excel_path}: {cleaned_sheets} sheets processed, "
                    f"{cleaned_headers} headers simplified, {cleaned_values} formatted values converted, "
                    f"{formatted_numbers} numbers formatted to {decimal_places} decimal places")

        except Exception as e:
            return f"Error cleaning coin market data: {str(e)}"

    @staticmethod
    def generate_coin_market_charts(source_excel_path, chart_excel_path, sheet_name=None,
                                    top_cryptocurrencies=30, include_fields=None):
        """
        Generates charts from coin market data and saves them to a new Excel file.
        Displays all percentage values as decimal numbers (not in percentage format).

        :param source_excel_path: Path to the source Excel file with coin market data
        :param chart_excel_path: Path to save the charts Excel file
        :param sheet_name: Name of the sheet to read from source (None for active sheet)
        :param top_cryptocurrencies: Number of top cryptocurrencies to include in charts
        :param include_fields: List of percent change fields to include (None for all available)
        :return: Status message
        """
        try:
            import openpyxl
            from openpyxl.chart import LineChart, BarChart, Reference
            from openpyxl.chart.axis import DateAxis
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            import datetime

            # Default percent change fields to chart if none specified
            if include_fields is None:
                include_fields = [
                    'percent_change_24h',
                    'percent_change_7d',
                    'percent_change_30d',
                    'percent_change_60d',
                    'percent_change_90d'
                ]

            # Load source workbook
            source_wb = openpyxl.load_workbook(source_excel_path)

            # Determine source sheet
            if sheet_name is None:
                # Use active sheet
                source_sheet = source_wb.active
                sheet_name = source_sheet.title
            else:
                # Use specified sheet
                if sheet_name not in source_wb.sheetnames:
                    return f"Error: Sheet '{sheet_name}' not found in {source_excel_path}"
                source_sheet = source_wb[sheet_name]

            # Create or load destination workbook for charts
            try:
                chart_wb = openpyxl.load_workbook(chart_excel_path)
            except FileNotFoundError:
                chart_wb = openpyxl.Workbook()
                # Remove default sheet
                if "Sheet" in chart_wb.sheetnames:
                    chart_wb.remove(chart_wb["Sheet"])

            # Create a new sheet in the chart workbook
            chart_sheet_name = f"{sheet_name}_charts"
            if chart_sheet_name in chart_wb.sheetnames:
                # Use existing sheet but clear it
                chart_sheet = chart_wb[chart_sheet_name]
                for row in chart_sheet.rows:
                    for cell in row:
                        cell.value = None
            else:
                # Create new sheet
                chart_sheet = chart_wb.create_sheet(title=chart_sheet_name)

            # Define styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Get headers from source sheet
            source_headers = []
            for cell in source_sheet[1]:
                source_headers.append(cell.value)

            # Find required column indices
            name_col_idx = None
            symbol_col_idx = None
            rank_col_idx = None
            market_cap_col_idx = None
            percent_change_col_indices = {}

            for i, header in enumerate(source_headers):
                if header == 'name':
                    name_col_idx = i
                elif header == 'symbol':
                    symbol_col_idx = i
                elif header == 'cmc_rank' or header == 'rank':
                    rank_col_idx = i
                elif header == 'market_cap':
                    market_cap_col_idx = i
                elif header in include_fields or (header.replace('quote_USD_', '') in include_fields):
                    clean_header = header.replace('quote_USD_', '')
                    if clean_header in include_fields:
                        percent_change_col_indices[clean_header] = i

            # Check if we found all the required columns
            if name_col_idx is None or symbol_col_idx is None:
                return "Error: Required columns 'name' and 'symbol' not found in source data"

            if not percent_change_col_indices:
                return f"Error: None of the required percent change columns {include_fields} found in source data"

            # Read and sort cryptocurrency data by rank or market cap
            crypto_data = []
            for row_idx, row in enumerate(source_sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip empty rows
                    continue

                crypto = {
                    'name': row[name_col_idx],
                    'symbol': row[symbol_col_idx],
                    'row_idx': row_idx
                }

                # Add rank if available
                if rank_col_idx is not None:
                    crypto['rank'] = row[rank_col_idx]

                # Add market cap if available
                if market_cap_col_idx is not None:
                    crypto['market_cap'] = row[market_cap_col_idx]

                # Add percent change values
                for field, col_idx in percent_change_col_indices.items():
                    if col_idx < len(row) and row[col_idx] is not None:
                        # Convert to float if not already
                        try:
                            crypto[field] = float(row[col_idx])
                        except (ValueError, TypeError):
                            crypto[field] = 0.0
                    else:
                        crypto[field] = 0.0

                crypto_data.append(crypto)

            # Sort by rank if available, otherwise by market cap if available
            if crypto_data:
                if all('rank' in crypto for crypto in crypto_data):
                    crypto_data.sort(key=lambda x: x.get('rank', float('inf')))
                elif all('market_cap' in crypto for crypto in crypto_data):
                    crypto_data.sort(key=lambda x: x.get('market_cap', 0), reverse=True)

            # Limit to top N cryptocurrencies
            crypto_data = crypto_data[:top_cryptocurrencies]

            # Get current timestamp for charts
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

            # Create summary table in chart sheet
            chart_sheet['A1'] = f"Cryptocurrency Percent Change Analysis - Generated: {timestamp}"
            chart_sheet['A1'].font = Font(bold=True, size=14)
            chart_sheet.merge_cells('A1:G1')
            chart_sheet['A1'].alignment = Alignment(horizontal='center')

            # Add headers for summary table
            chart_sheet['A3'] = "Rank"
            chart_sheet['B3'] = "Symbol"
            chart_sheet['C3'] = "Name"

            col_offset = 3
            for field in include_fields:
                # Clean up the headers for display
                display_name = field.replace('percent_change_', '')
                # Format the header based on time period
                if display_name == '24h':
                    header_text = "24 Hours"
                elif display_name == '7d':
                    header_text = "7 Days"
                elif display_name == '30d':
                    header_text = "30 Days"
                elif display_name == '60d':
                    header_text = "60 Days"
                elif display_name == '90d':
                    header_text = "90 Days"
                else:
                    header_text = display_name.upper()

                chart_sheet.cell(row=3, column=col_offset+1, value=header_text)
                col_offset += 1

            # Style headers
            for col in range(1, col_offset+1):
                cell = chart_sheet.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Add data rows
            for i, crypto in enumerate(crypto_data, start=4):
                chart_sheet.cell(row=i, column=1, value=i-3)  # Rank in chart
                chart_sheet.cell(row=i, column=2, value=crypto['symbol'])
                chart_sheet.cell(row=i, column=3, value=crypto['name'])

                col_offset = 3
                for field in include_fields:
                    value = crypto.get(field, 0)
                    cell = chart_sheet.cell(row=i, column=col_offset+1, value=value)
                    # Format as decimal with 4 decimal places (no percentage)
                    cell.number_format = '0.0000'
                    # Color-code cells: green for positive, red for negative
                    if value > 0:
                        cell.font = Font(color="006100")  # Dark green
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value < 0:
                        cell.font = Font(color="9C0006")  # Dark red
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    col_offset += 1

                # Add borders to all cells in this row
                for col in range(1, col_offset+1):
                    chart_sheet.cell(row=i, column=col).border = border

            # Auto-adjust column widths
            for col in range(1, col_offset+1):
                max_length = 0
                column_letter = get_column_letter(col)

                for row in range(1, len(crypto_data) + 5):
                    cell = chart_sheet.cell(row=row, column=col)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, min(cell_length, 30))

                adjusted_width = max_length + 4
                chart_sheet.column_dimensions[column_letter].width = adjusted_width

            # Create charts - one chart for each time period
            chart_row_start = len(crypto_data) + 6  # Start after the table

            for i, field in enumerate(include_fields):
                # Clean field name for chart title
                display_name = field.replace('percent_change_', '')
                if display_name == '24h':
                    title_text = "24 Hours Change"
                elif display_name == '7d':
                    title_text = "7 Days Change"
                elif display_name == '30d':
                    title_text = "30 Days Change"
                elif display_name == '60d':
                    title_text = "60 Days Change"
                elif display_name == '90d':
                    title_text = "90 Days Change"
                else:
                    title_text = f"{display_name.upper()} Change"

                # Create chart for this time period
                chart = BarChart()
                chart.type = "col"  # Column chart
                chart.style = 10
                chart.title = title_text
                chart.y_axis.title = "Decimal Change"  # Not "Percent Change"
                chart.x_axis.title = "Cryptocurrency"

                # Add data for this chart
                data_col = 4 + i  # Column containing this percent change data
                data = Reference(chart_sheet, min_col=data_col, min_row=3, max_row=len(crypto_data) + 3, max_col=data_col)
                cats = Reference(chart_sheet, min_col=2, min_row=4, max_row=len(crypto_data) + 3, max_col=2)  # Symbol column

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)

                # Set chart size and position
                chart.height = 15  # Height in cm
                chart.width = 20   # Width in cm

                # Add chart to sheet, leave 2 row spacing between charts
                chart_position = f"A{chart_row_start + (i * 20)}"
                chart_sheet.add_chart(chart, chart_position)

            # Create a comparison line chart for all time periods
            comparison_chart = LineChart()
            comparison_chart.title = "Comparison of All Change Periods"  # Not "Percent Change"
            comparison_chart.style = 12
            comparison_chart.y_axis.title = "Decimal Change"  # Not "Percent Change"
            comparison_chart.x_axis.title = "Cryptocurrency"
            comparison_chart.height = 15
            comparison_chart.width = 25

            # Add data series for each time period
            for i, field in enumerate(include_fields):
                data_col = 4 + i
                data = Reference(chart_sheet, min_col=data_col, min_row=3, max_row=len(crypto_data) + 3, max_col=data_col)
                if i == 0:
                    comparison_chart.add_data(data, titles_from_data=True)
                else:
                    comparison_chart.add_data(data, titles_from_data=True)

            cats = Reference(chart_sheet, min_col=2, min_row=4, max_row=len(crypto_data) + 3, max_col=2)
            comparison_chart.set_categories(cats)

            # Add comparison chart to the end
            chart_position = f"A{chart_row_start + (len(include_fields) * 20)}"
            chart_sheet.add_chart(comparison_chart, chart_position)

            # Create explanation note
            note_row = chart_row_start + (len(include_fields) * 20) + 20
            chart_sheet.cell(row=note_row, column=1, value="Note: All values are displayed as decimal numbers (not percentages).")
            chart_sheet.cell(row=note_row, column=1).font = Font(italic=True)
            chart_sheet.merge_cells(f'A{note_row}:E{note_row}')

            # Save chart workbook
            chart_wb.save(chart_excel_path)

            return f"Successfully generated charts for {len(crypto_data)} cryptocurrencies with {len(include_fields)} change metrics in {chart_excel_path}"

        except Exception as e:
            return f"Error generating coin market charts: {str(e)}"